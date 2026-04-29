#!/usr/bin/env python3
"""
빈박스 + 실배송 엑셀 파싱 통합 스크립트 (v1.1.0)

- 빈박스(체험단): 3pl N/C 체험단 엑셀 → 주문번호 + CJ송장번호
- 실배송(풀필먼트): 발주조회_*.xlsx (암호화) → 주문번호 + 롯데송장번호
- 자동 시나리오 감지 또는 --scenario 명시 가능
"""

import openpyxl
import json
import os
import re
import sys
import argparse
import tempfile

try:
    import msoffcrypto
    HAS_MSOFFCRYPTO = True
except ImportError:
    HAS_MSOFFCRYPTO = False


def normalize_waybill(wybl_str):
    """송장번호 정규화: 하이픈 제거, 숫자만 추출"""
    if not wybl_str:
        return ''
    return re.sub(r'[^0-9]', '', str(wybl_str).strip())


def detect_mall(filename):
    """파일명으로 쇼핑몰 판별 (빈박스용)"""
    fname = os.path.basename(filename).lower()
    if '3pl n' in fname or 'n 체험단' in fname:
        return 'naver'
    elif '3pl c' in fname or 'c 체험단' in fname:
        return 'coupang'
    return 'unknown'


def detect_scenario(filepath):
    """파일명·헤더로 시나리오 자동 감지: 'binbox' | 'realship'"""
    base = os.path.basename(filepath).lower()
    if base.startswith('발주조회') or 'orders_' in base or 'fulfillment' in base:
        return 'realship'
    if '체험단' in base or '3pl' in base:
        return 'binbox'
    # 헤더 fallback (암호화 파일은 못 읽으므로 binbox 우선)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(h or '') for h in first_row]
        if any('CJ대한통운' in h or 'CJ 대한통운' in h for h in headers):
            return 'binbox'
        if any('오더코드' in h for h in headers) and any('판매상품명' in h for h in headers):
            return 'realship'
    except Exception:
        pass
    return 'unknown'


def decrypt_xlsx(src_path, password):
    """암호화된 .xlsx 복호화 → 임시 복호화 파일 경로 반환"""
    if not HAS_MSOFFCRYPTO:
        raise RuntimeError("msoffcrypto-tool 미설치. pip install msoffcrypto-tool --break-system-packages")
    fd, tmp = tempfile.mkstemp(suffix='.xlsx', prefix='dec_')
    os.close(fd)
    with open(src_path, 'rb') as fin:
        ole = msoffcrypto.OfficeFile(fin)
        ole.load_key(password=password)
        with open(tmp, 'wb') as fout:
            ole.decrypt(fout)
    return tmp


# ============================================================
# 빈박스 파서
# ============================================================
def parse_binbox(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in header_row]

    col_map = {}
    for idx, h in enumerate(headers):
        h_lower = h.lower().replace(' ', '')
        if '주문번호' in h:
            col_map['ordNo'] = idx
        elif 'cj' in h_lower and '송장' in h:
            col_map['wyblNo'] = idx
        elif h == '체험단 이름':
            col_map['name'] = idx
        elif h == '체험단 전화번호':
            col_map['phone'] = idx
        elif h == '체험단 주소':
            col_map['address'] = idx
        elif '상품 메모' in h or '상품메모' in h:
            col_map['product'] = idx
        elif '선택 옵션' in h or '선택옵션' in h:
            col_map['option'] = idx

    mall = detect_mall(filepath)
    fname = os.path.basename(filepath)
    orders = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals or all(v is None for v in vals):
            continue

        ord_no = str(vals[col_map.get('ordNo', 0)]).strip() if col_map.get('ordNo') is not None else ''
        if not ord_no or ord_no == 'None':
            continue

        wybl_raw = str(vals[col_map.get('wyblNo', -1)]) if col_map.get('wyblNo') is not None and col_map['wyblNo'] < len(vals) else ''
        wybl_no = normalize_waybill(wybl_raw)

        def get(k):
            ci = col_map.get(k)
            if ci is None or ci >= len(vals) or vals[ci] is None:
                return ''
            return str(vals[ci]).strip()

        orders.append({
            'shmaOrdNo': ord_no,
            'wyblNo': wybl_no,
            'wyblRaw': wybl_raw.strip() if wybl_raw != 'None' else '',
            'name': get('name'),
            'phone': get('phone'),
            'address': get('address'),
            'product': get('product'),
            'option': get('option'),
            'file': fname,
            'mall': mall,
            'scenario': 'binbox',
        })

    wb.close()
    return orders


# ============================================================
# 실배송 파서 (발주조회 양식)
# ============================================================
def parse_realship(filepath, password=None):
    """발주조회 엑셀 파싱. 암호화되어 있으면 password로 복호화."""
    src_for_open = filepath
    tmp_dec = None
    try:
        # 암호화 여부 확인
        with open(filepath, 'rb') as fin:
            head = fin.read(8)
        is_encrypted = head[:4] == b'\xD0\xCF\x11\xE0'  # CDFV2 magic
        if is_encrypted:
            if not password:
                raise RuntimeError(f"암호화된 파일이지만 비밀번호 미제공: {filepath}")
            tmp_dec = decrypt_xlsx(filepath, password)
            src_for_open = tmp_dec

        wb = openpyxl.load_workbook(src_for_open, data_only=True)
        # 시트명 "발주조회" 우선 사용, 없으면 active
        ws = wb['발주조회'] if '발주조회' in wb.sheetnames else wb.active
        fname = os.path.basename(filepath)
        orders = []

        # 발주조회 양식의 고정 컬럼 위치
        # 1: 오더코드, 2: 릴리즈코드, 3: 택배사, 4: 송장번호, 5: 회사명, 6: 고유코드
        # 7: 판매상품명, 8: 수량, 11: 받는분, 12: 전화번호1, 14: 우편번호, 15: 주소
        # 16: 배송메시지, 17: 주문번호
        miss_idx = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if not vals or all(v is None for v in vals):
                continue
            if len(vals) < 17:
                continue

            wybl_raw = vals[3] if len(vals) > 3 else ''
            company = vals[4] if len(vals) > 4 else ''
            product = vals[6] if len(vals) > 6 else ''
            qty = vals[7] if len(vals) > 7 else 0
            recvr = vals[10] if len(vals) > 10 else ''
            phone = vals[11] if len(vals) > 11 else ''
            addr = vals[14] if len(vals) > 14 else ''
            msg = vals[15] if len(vals) > 15 else ''
            ordno = vals[16] if len(vals) > 16 else ''
            courier = vals[2] if len(vals) > 2 else ''

            wybl_no = normalize_waybill(wybl_raw)

            # 주문번호 비어있으면 누락 발송 케이스
            if not ordno or str(ordno).strip() in ('', 'None'):
                miss_idx += 1
                shma_ord_no = f"MISS_{wybl_no or miss_idx}"
            else:
                shma_ord_no = str(ordno).strip()

            orders.append({
                'shmaOrdNo': shma_ord_no,
                'wyblNo': wybl_no,
                'wyblRaw': str(wybl_raw or '').strip(),
                'courier': str(courier or '').strip(),
                'company': str(company or '').strip(),
                'product': str(product or '').strip(),
                'qty': qty if isinstance(qty, (int, float)) else 0,
                'name': str(recvr or '').strip(),
                'phone': str(phone or '').strip(),
                'address': str(addr or '').strip(),
                'msg': str(msg or '').strip(),
                'file': fname,
                'scenario': 'realship',
            })

        wb.close()
        return orders
    finally:
        if tmp_dec and os.path.exists(tmp_dec):
            try:
                os.unlink(tmp_dec)
            except Exception:
                pass


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='빈박스/실배송 엑셀 통합 파싱')
    parser.add_argument('--input-dir', help='엑셀 파일 디렉토리')
    parser.add_argument('--files', nargs='+', help='개별 엑셀 파일 경로들')
    parser.add_argument('--output', required=True, help='출력 JSON 경로')
    parser.add_argument('--scenario', choices=['binbox', 'realship', 'auto'], default='auto')
    parser.add_argument('--password', help='실배송 발주조회 엑셀 비밀번호')
    args = parser.parse_args()

    files_to_parse = []
    if args.files:
        files_to_parse = args.files
    elif args.input_dir:
        for f in os.listdir(args.input_dir):
            if not f.endswith('.xlsx'):
                continue
            fl = f.lower()
            # 시나리오에 맞는 파일만 선별 (auto면 둘 다)
            if args.scenario in ('binbox', 'auto') and ('3pl' in fl or '체험단' in f):
                files_to_parse.append(os.path.join(args.input_dir, f))
                continue
            if args.scenario in ('realship', 'auto') and (f.startswith('발주조회') or 'orders_' in fl):
                files_to_parse.append(os.path.join(args.input_dir, f))
                continue
    else:
        print("ERROR: --input-dir 또는 --files 중 하나를 지정", file=sys.stderr)
        sys.exit(1)

    if not files_to_parse:
        print("ERROR: 파싱할 엑셀이 없음", file=sys.stderr)
        sys.exit(1)

    all_orders = []
    by_file = {}
    errors = []
    detected_scenarios = set()

    for fpath in sorted(files_to_parse):
        try:
            scn = args.scenario if args.scenario != 'auto' else detect_scenario(fpath)
            if scn == 'binbox':
                orders = parse_binbox(fpath)
            elif scn == 'realship':
                orders = parse_realship(fpath, password=args.password)
            else:
                errors.append({'file': os.path.basename(fpath), 'error': f'unknown scenario'})
                continue
            detected_scenarios.add(scn)
            all_orders.extend(orders)
            by_file[os.path.basename(fpath)] = len(orders)
        except Exception as e:
            errors.append({'file': os.path.basename(fpath), 'error': str(e)})

    no_wybl = [o for o in all_orders if not o['wyblNo']]
    miss = [o for o in all_orders if str(o.get('shmaOrdNo', '')).startswith('MISS_')]

    result = {
        'scenario': list(detected_scenarios),
        'total': len(all_orders),
        'orders': all_orders,
        'by_file': by_file,
        'no_waybill_count': len(no_wybl),
        'miss_count': len(miss),
        'errors': errors,
    }

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"시나리오: {list(detected_scenarios)}")
    print(f"총 {len(all_orders)}건 추출 ({len(files_to_parse)}개 파일)")
    for fname, cnt in by_file.items():
        print(f"  {fname}: {cnt}건")
    if no_wybl:
        print(f"⚠️ 송장번호 없는 주문: {len(no_wybl)}건")
    if miss:
        print(f"⚠️ 누락 발송(MISS_) 케이스: {len(miss)}건")
    if errors:
        print(f"❌ 오류 파일: {len(errors)}개")
        for e in errors:
            print(f"  {e['file']}: {e['error']}")


if __name__ == '__main__':
    main()
